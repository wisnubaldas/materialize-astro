import logging
from io import BytesIO

import requests
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from requests.auth import HTTPBasicAuth
from sqlalchemy import and_, case, func, literal_column, text
from sqlalchemy.orm import Session
from sqlalchemy.sql.elements import ColumnElement

from app.db.mysql import SessionDB1R, SessionDB2R
from app.models.BaseDB1.hubnet_request import HubnetRequest
from app.repository.hubnet_request_repository import HubnetRequestRepository
from app.schemas.datatables_schema import DataTablesParams, DataTablesResponse
from app.schemas.delete_data_terkirim_schema import DeleteDataTerkirimSchema
from app.schemas.hubnet_request_schema import HubnetRequestGet
from app.services.datatables_service import DataTablesService
from app.utils.env import ENV
from app.utils.helper import HELPER

logger = logging.getLogger("hubnet")


class HbnetRequestService:
    def __init__(self, repo: HubnetRequestRepository):
        self.repo = repo

    def data_sending_per_bulan(self, bulan: str) -> list[dict[str, int | str]]:
        return self.repo.get_data_sending_perbulan(bulan)

    def get_data_export_excel(self, bulan: str) -> list[HubnetRequest]:
        _data = self.repo.export_to_excel(bulan)
        return _data

    @staticmethod
    def get_data_request(
        db: Session, params: DataTablesParams
    ) -> DataTablesResponse[HubnetRequestGet]:
        data_request = DataTablesService(
            model=HubnetRequest,
            schema=HubnetRequestGet,
            search_columns=["AWB_NO", "FLT_DATE", "FLT_NUMBER", "IS_INTERNATIONAL", "IS_EKSPOR"],
            custom_filters=["AWB_NO", "FLT_DATE", "FLT_NUMBER", "IS_INTERNATIONAL", "IS_EKSPOR"],
        )
        return data_request.get_datatable(db=db, params=params)

    # upload Excel manifest di insert ke table request HUBNET
    @staticmethod
    def upload_manifest(file: UploadFile, db: Session):  # noqa: PLR0912, PLR0915
        # validasi ekstensi
        if not file.filename.endswith((".xlsx", ".xlsm")):  # type: ignore
            raise HTTPException(
                status_code=400, detail="Format file tidak valid, gunakan Excel (.xlsx / .xlsm)"
            )

        contents = file.file.read()
        wb = load_workbook(filename=BytesIO(contents), read_only=True)
        ws = wb.active

        # validasi header
        expected_headers = [
            "AWB_NO",
            "FLT_NUMBER",
            "FLT_DATE",
            "ORI",
            "DEST",
            "T",
            "K",
            "CH_WEIGHT",
            "MC",
            "KATEGORI_CARGO",
        ]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]  # pyright: ignore[reportOptionalMemberAccess]
        if headers != expected_headers:
            logger.error(f"Header Excel tidak sesuai {file.filename}")
            raise HTTPException(
                status_code=400,
                detail={
                    "message": "Header file tidak sesuai !",
                    "error": {"header_file_upload": headers, "header_file_valid": expected_headers},
                },
            )

        batch = []
        batch_size = 500
        no_customer = 0

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):  # pyright: ignore[reportOptionalMemberAccess]
            if not any(row):
                continue
            AWB_NO, FLT_NUMBER, FLT_DATE, ORI, DEST, T, K, CH_WEIGHT, MC, KATEGORI_CARGO = row
            # validasi kolom wajib
            if (
                not AWB_NO
                or not FLT_NUMBER
                or not FLT_DATE
                or not ORI
                or not DEST
                or not T
                or not K
                or not MC
                or not KATEGORI_CARGO
            ):
                logger.error(f"Data tidak lengkap di baris {idx}")
                raise HTTPException(
                    status_code=400,
                    detail={
                        "message": f"Data tidak lengkap di baris {idx}",
                        "error": {
                            "header_file_upload": headers,
                            "header_file_valid": expected_headers,
                        },
                    },
                )

            customer = None
            category = None
            # mapping kategori cargo ke IS_INTERNATIONAL dan IS_EKSPOR
            # cari data customer nya
            if KATEGORI_CARGO == "EKSPORT":
                category = "EKSPORT"
                IS_INTERNATIONAL, IS_EKSPOR = 1, 1
                customer = HbnetRequestService.__get_hostawb(
                    awb=AWB_NO, qfile="app/repository/hubnet_query/get_ekspor_hawb.sql"
                )
                REMARKS = customer.get("descriptiongoods") if customer else None
                AGT_NAME = customer.get("agenCompany") if customer else None
                SHP_ADD = customer.get("shipperAddr") if customer else None
                SHP_NAME = customer.get("shipperName") if customer else None
                CNE_NAME = customer.get("conssigneeName") if customer else None
                CNE_ADD = customer.get("ConsigneeAddr") if customer else None

            elif KATEGORI_CARGO == "IMPORT":
                category = "IMPORT"
                IS_INTERNATIONAL, IS_EKSPOR = 1, 0
                customer = HbnetRequestService.__get_hostawb(
                    awb=AWB_NO, qfile="app/repository/hubnet_query/get_imp_hostawb.sql"
                )
                REMARKS = customer.get("DescriptionGoods") if customer else None
                AGT_NAME = customer.get("AgenCode") if customer else None
                SHP_ADD = customer.get("shipperAddr") if customer else None
                SHP_NAME = customer.get("shippername") if customer else None
                CNE_NAME = customer.get("Consigneename") if customer else None
                CNE_ADD = customer.get("consigneeAddr") if customer else None
            elif KATEGORI_CARGO == "OUTGOING":
                category = "OUTGOING"

                IS_INTERNATIONAL, IS_EKSPOR = 0, 1
                customer = HbnetRequestService.__get_hostawb(
                    awb=AWB_NO, qfile="app/repository/hubnet_query/get_outgoing_customer.sql"
                )
                REMARKS = customer.get("kindofgood") if customer else None
                AGT_NAME = customer.get("agenName") if customer else None
                SHP_NAME = customer.get("shipperName") if customer else None
                SHP_ADD = customer.get("shipperAddr") if customer else None
                CNE_NAME = customer.get("consigneeName") if customer else None
                CNE_ADD = customer.get("consigneeAddr") if customer else None

            elif KATEGORI_CARGO == "INCOMING":
                category = "INCOMING"
                pass
                # IS_INTERNATIONAL, IS_EKSPOR = 0, 0
                # customer = HbnetRequestService.__get_hostawb(
                #     awb=AWB_NO, qfile="app/repository/query/get_imp_hostawb.sql"
                # )
                # REMARKS = customer.get("descriptiongoods") if customer else None
                # AGT_NAME = customer.get("AgenCode") if customer else None
                # SHP_ADD = customer.get("shipperaddress") if customer else None
                # SHP_NAME = customer.get("shippername") if customer else None
                # CNE_NAME = customer.get("Consigneename") if customer else None
                # CNE_ADD = customer.get("Consigneeaddress") if customer else None
            else:
                raise HTTPException(
                    status_code=400,
                    detail={
                        "message": f"Kategori cargo tidak valid di baris {idx}: {KATEGORI_CARGO}",
                        "error": {
                            "header_file_upload": headers,
                            "header_file_valid": expected_headers,
                        },
                    },
                )
                # tambahin jam ke FLT_DATE
            flt_datetime = HbnetRequestService.__combine_date_with_current_time(FLT_DATE)
            if customer is not None:
                batch.append(
                    HubnetRequest(
                        AWB_NO=AWB_NO,
                        FLT_NUMBER=FLT_NUMBER,
                        FLT_DATE=flt_datetime,
                        ORI=ORI,
                        DEST=DEST,
                        T=T,
                        K=K,
                        CH_WEIGHT=CH_WEIGHT,
                        MC=MC,
                        IS_INTERNATIONAL=IS_INTERNATIONAL,
                        IS_EKSPOR=IS_EKSPOR,
                        FLT_NUMBER1=FLT_NUMBER,
                        FLT_DATE1=flt_datetime,
                        ORI1=ORI,
                        AGT_NAME=AGT_NAME,
                        AGT_ADD="",
                        SHP_ADD=SHP_ADD,
                        SHP_NAME=SHP_NAME,
                        CNE_NAME=CNE_NAME,
                        CNE_ADD=CNE_ADD,
                        KATEGORI_CARGO=KATEGORI_CARGO,
                        COMMODITY=REMARKS,
                        CARGO_TREATMENT="",
                        REMARKS=REMARKS,
                    )
                )
                logger.info(f"host insert {AWB_NO}")
            else:
                no_customer = no_customer + 1
                logger.error(f"Data custemer tidak ada di {category} host: {AWB_NO}")
            if len(batch) == batch_size:
                db.bulk_save_objects(batch)
                db.commit()
                batch = []
                logger.info(f"sukses insert upload {len(batch)} data")
        # save sisa dari batch
        if batch:
            db.bulk_save_objects(batch)
            db.commit()

        return {"message": f"Upload berhasil, data tidak lengkap {no_customer} data"}

    # nampilin data terakhir ke send di front
    @staticmethod
    def last_sending():
        session = SessionDB1R()
        data = (
            session.query(HubnetRequest)
            .filter(HubnetRequest.IS_SEND == "1")
            .order_by(HubnetRequest.created_at.desc())
            .first()
        )
        session.close()
        return data

    # nampiling data terkirim dari API HUBNET
    @staticmethod
    def get_data_terkirim(flt_date: str, page: int = 1, per_page: int = 10):
        try:
            payload = {"FLT_DATE": flt_date}
            logger.info(
                "Mengambil data terkirim Hubnet",
                extra={
                    "event": "hubnet.fetch.start",
                    "flight_date": flt_date,
                    "page": page,
                    "per_page": per_page,
                },
            )
            response = requests.post(
                f"{ENV.HUBNET_URL}/nle-udara/get-data-logistik?page={page}&per_page={per_page}",
                data=payload,
                auth=HTTPBasicAuth(ENV.HUBNET_USER, ENV.HUBNET_PASSWORD),
            )
            response.raise_for_status()
            logger.info(
                "Menampilkan data terkirim dari Hubnet",
                extra={
                    "event": "hubnet.fetch.success",
                    "status": response.status_code,
                    "flight_date": flt_date,
                },
            )
            return response.json()
        except requests.exceptions.RequestException as e:
            response_status = getattr(e.response, "status_code", None)
            status_code = response_status or 502
            logger.exception(
                "Terjadi kesalahan saat mengambil data Hubnet",
                extra={"event": "hubnet.fetch.error", "status": status_code},
            )
            response_text = getattr(e.response, "text", None)
            if response_text:
                logger.error(
                    "Respons server Hubnet: %s",
                    response_text,
                    extra={"event": "hubnet.fetch.response_body"},
                )
            error_detail = response_text or str(e)
            raise HTTPException(
                status_code=status_code,
                detail={
                    "message": "Gagal terhubung ke layanan Hubnet.",
                    "error": error_detail,
                },
            ) from e

    # hapus data terkerim di hubnet API
    @staticmethod
    def delete_data_terkirim_api(params: DeleteDataTerkirimSchema):
        try:
            # Konversi setiap model Pydantic di dalam list menjadi dict
            payload = [item.model_dump() for item in params]

            response = requests.post(
                f"{ENV.HUBNET_URL}/nle-udara/delete-data-logistik",
                json=payload,
                auth=HTTPBasicAuth(ENV.HUBNET_USER, ENV.HUBNET_PASSWORD),
            )
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            status_code = getattr(e.response, "status_code", None)
            logger.exception(
                "Terjadi kesalahan saat menghapus data Hubnet",
                extra={"event": "hubnet.delete.error", "status": status_code},
            )
            if getattr(e.response, "text", None):
                logger.error(
                    "Respons server Hubnet (delete): %s",
                    e.response.text,
                    extra={"event": "hubnet.delete.response_body"},
                )
            return None

    @staticmethod
    def __combine_date_with_current_time(FLT_DATE, idx=0):  # noqa: N803
        try:
            # # --- Parsing FLT_DATE ---
            # if isinstance(FLT_DATE, datetime):
            #     base_date = FLT_DATE
            # else:
            #     # coba 2 format umum
            #     try:
            #         base_date = datetime.strptime(str(FLT_DATE), "%d-%m-%Y")
            #     except ValueError:
            #         base_date = datetime.strptime(str(FLT_DATE), "%Y-%m-%d")

            # # --- Tambahkan jam saat ini (Asia/Jakarta) ---
            # tz = pytz.timezone("Asia/Jakarta")
            # now = datetime.now(tz)
            # flt_datetime = base_date.replace(
            #     hour=now.hour, minute=now.minute, second=now.second, microsecond=0
            # )

            # # pastikan timezone-aware
            # flt_datetime = tz.localize(flt_datetime)
            # return flt_datetime
            return FLT_DATE

        except Exception:
            logger.error(f"Baris {idx}: format tanggal tidak valid ({FLT_DATE})")
            raise ValueError(f"Baris {idx}: format tanggal tidak valid ({FLT_DATE})")  # noqa: B904

    @staticmethod
    def __get_hostawb(awb: str, qfile: str):
        try:
            db2 = SessionDB2R()
            query = HELPER.load_sql_query(qfile)
            param = {"awb": awb}
            sql = text(query)
            customers = db2.execute(sql, param).mappings().first()
        except Exception as e:
            logger.exception(
                "Gagal mengambil data host AWB untuk customer",
                extra={
                    "event": "hubnet.hostawb.error",
                    "awb": awb,
                    "query_file": qfile,
                    "error_message": str(e),  # Menampilkan pesan error
                    "executed_sql": str(sql),  # Menampilkan query SQL yang dieksekusi
                    "parameters": param,  # Menampilkan parameter yang digunakan
                },
            )
        finally:
            db2.close()

        return customers

    @staticmethod
    def _create_conditional_count_expression(
        conditions: list[ColumnElement], label: str
    ) -> ColumnElement:
        combined_condition = and_(*conditions)
        count_expression = func.sum(
            case(
                (combined_condition, literal_column("1")),  # Jika kriteria benar, hitung 1
                else_=literal_column("0"),  # Jika kriteria salah, hitung 0
            )
        ).label(label)
        return count_expression
