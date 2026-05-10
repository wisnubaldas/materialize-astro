import json
import logging
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

from fastapi import HTTPException
from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy.exc import IntegrityError, ProgrammingError
from sqlalchemy.orm import Session

from app.models.BaseDB1.build_up_dead_stock import BuildUpDeadStock
from app.models.BaseDB1.build_up_detail import BuildUpDetail
from app.models.BaseDB1.build_up_header import BuildUpHeader
from app.utils.helper import PDF_DIR

logger = logging.getLogger("warehouse")

FLIGHT_HEADERS = [
    "airline_code",
    "flight_number",
    "flight_date",
    "aircraft_registration",
    "point_of_loading",
    "point_of_unloading",
    "total_pieces",
    "total_weight_kg",
    "total_volume",
    "source_document",
    "raw_text",
]

FLIGHT_HEADERS_WITHOUT_TOTAL_VOLUME = [
    "airline_code",
    "flight_number",
    "flight_date",
    "aircraft_registration",
    "point_of_loading",
    "point_of_unloading",
    "total_pieces",
    "total_weight_kg",
    "source_document",
    "raw_text",
]

ULD_HEADERS = [
    "flight_number",
    "flight_date",
    "uld_type",
    "uld_number",
    "uld_owner",
    "destination",
    "remarks",
]

MAWB_HEADERS = [
    "flight_number",
    "flight_date",
    "uld_type",
    "uld_number",
    "mawb_prefix",
    "mawb_number",
    "pieces",
    "total_pieces",
    "total_weight_kg",
    "weight_kg",
    "volume",
    "nature_of_goods",
    "route",
    "transit_flag",
]

MAWB_HEADERS_WITHOUT_TOTAL_WEIGHT = [
    "flight_number",
    "flight_date",
    "uld_type",
    "uld_number",
    "mawb_prefix",
    "mawb_number",
    "pieces",
    "total_pieces",
    "weight_kg",
    "volume",
    "nature_of_goods",
    "route",
    "transit_flag",
]

MAWB_HEADERS_LEGACY = [
    "flight_number",
    "flight_date",
    "uld_type",
    "uld_number",
    "mawb_prefix",
    "mawb_number",
    "pieces",
    "weight_kg",
    "volume",
    "nature_of_goods",
    "route",
    "transit_flag",
]


def _normalize_headers(raw_headers: tuple) -> list[str | None]:
    headers: list[str | None] = []
    for header in raw_headers:
        if header is None:
            headers.append(None)
            continue
        if isinstance(header, str):
            headers.append(header.strip())
        else:
            headers.append(str(header).strip())
    while headers and headers[-1] is None:
        headers.pop()
    return headers


def _row_has_values(row: tuple) -> bool:
    for value in row:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return True
    return False


def _normalize_identifier(value, pad: int | None = None) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        result = value.strip()
    elif isinstance(value, float):
        result = str(int(value)) if value.is_integer() else str(value)
    else:
        result = str(value).strip()
    if not result:
        return None
    if pad:
        return result.zfill(pad)
    return result


def _normalize_awb_key(value: str | None) -> str:
    if not value:
        return ""
    return value.strip().upper()


def _is_missing_dead_stock_table_error(exc: ProgrammingError) -> bool:
    original = getattr(exc, "orig", None)
    if original is not None:
        args = getattr(original, "args", ())
        if args and args[0] == 1146:
            return True
    message = str(exc).lower()
    return "build_up_dead_stock" in message and "doesn't exist" in message


def _parse_date(value, field: str, row_idx: int) -> date:
    if value is None or value == "":
        raise HTTPException(status_code=400, detail=f"{field} kosong di baris {row_idx}.")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            parsed = from_excel(value)
            return parsed.date() if isinstance(parsed, datetime) else parsed
        except Exception as exc:
            raise HTTPException(
                status_code=400, detail=f"Format tanggal {field} tidak valid di baris {row_idx}."
            ) from exc
    if isinstance(value, str):
        raw = value.strip()
        for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(raw, fmt).replace(tzinfo=timezone.utc).date()
            except ValueError:
                continue
    raise HTTPException(
        status_code=400, detail=f"Format tanggal {field} tidak valid di baris {row_idx}."
    )


def _to_int(value, field: str, row_idx: int, default: int = 0) -> int:
    if value in (None, ""):
        return default
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise HTTPException(
            status_code=400, detail=f"Format {field} tidak valid di baris {row_idx}."
        ) from exc


def _to_decimal(value, field: str, row_idx: int, default: Decimal = Decimal("0")) -> Decimal:
    if value in (None, ""):
        return default
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise HTTPException(
            status_code=400, detail=f"Format {field} tidak valid di baris {row_idx}."
        ) from exc


def _to_bool(value) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return int(value) != 0
    if isinstance(value, str):
        return value.strip().lower() in ("1", "true", "yes", "y")
    return False


def _read_headers(ws, expected: list[str], sheet_name: str) -> list[str | None]:
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status_code=400, detail=f"Sheet {sheet_name} tidak memiliki header.")
    headers = _normalize_headers(header_row)
    if headers != expected:
        raise HTTPException(
            status_code=400,
            detail={
                "message": f"Header sheet {sheet_name} tidak sesuai.",
                "error": {"header_file_upload": headers, "header_file_valid": expected},
            },
        )
    return headers


def _read_headers_any(ws, expected_groups: list[list[str]], sheet_name: str) -> list[str | None]:
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status_code=400, detail=f"Sheet {sheet_name} tidak memiliki header.")
    headers = _normalize_headers(header_row)
    for expected in expected_groups:
        if headers == expected:
            return headers
    raise HTTPException(
        status_code=400,
        detail={
            "message": f"Header sheet {sheet_name} tidak sesuai.",
            "error": {"header_file_upload": headers, "header_file_valid": expected_groups},
        },
    )


def _format_decimal(value: Decimal) -> str:
    try:
        return f"{value:.2f}"
    except Exception:
        return str(value)


def _compose_detail_remark(
    base_remark: str | None,
    transit_marker: str | None,
) -> str | None:
    parts: list[str] = []
    cleaned_base = (base_remark or "").strip()
    cleaned_transit = (transit_marker or "").strip()
    if cleaned_base:
        parts.append(cleaned_base)
    if cleaned_transit:
        parts.append(cleaned_transit)
    return " | ".join(parts) if parts else None


class BuildupService:
    @staticmethod
    def _build_manifest_payload(flights: list[dict]) -> dict | None:
        if not flights:
            return None

        first = flights[0]
        grouped_ulds: dict[tuple[str, str, str], dict] = {}
        ordered_ulds: list[dict] = []

        for flight in flights:
            for uld in flight.get("ulds", []):
                uld_type = str(uld.get("uld_type") or "").strip()
                uld_number = str(uld.get("uld_number") or "").strip()
                uld_owner = str(uld.get("uld_owner") or "").strip()
                key = (uld_type, uld_number, uld_owner)

                existing = grouped_ulds.get(key)
                if existing is None:
                    existing = {
                        "uld_type": uld.get("uld_type"),
                        "uld_number": uld.get("uld_number"),
                        "uld_owner": uld.get("uld_owner"),
                        "destination": uld.get("destination"),
                        "remarks": uld.get("remarks"),
                        "mawbs": [],
                    }
                    grouped_ulds[key] = existing
                    ordered_ulds.append(existing)
                else:
                    if not existing.get("destination") and uld.get("destination"):
                        existing["destination"] = uld.get("destination")
                    if not existing.get("remarks") and uld.get("remarks"):
                        existing["remarks"] = uld.get("remarks")

                existing["mawbs"].extend(uld.get("mawbs", []))

        return {
            "airline_code": first.get("airline_code"),
            "flight_number": first.get("flight_number"),
            "flight_date": first.get("flight_date"),
            "aircraft_registration": first.get("aircraft_registration"),
            "point_of_loading": first.get("point_of_loading"),
            "point_of_unloading": first.get("point_of_unloading"),
            "ulds": ordered_ulds,
        }

    @staticmethod
    def _build_workbook_from_payload(payload_json: str) -> bytes:
        try:
            payload = json.loads(payload_json)
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=400, detail="Format payload_json tidak valid.") from exc

        if not isinstance(payload, dict):
            raise HTTPException(status_code=400, detail="payload_json harus berupa object.")

        flight_rows = payload.get("flight_manifest", [])
        uld_rows = payload.get("uld", [])
        mawb_rows = payload.get("mawb", [])
        if (
            not isinstance(flight_rows, list)
            or not isinstance(uld_rows, list)
            or not isinstance(mawb_rows, list)
        ):
            raise HTTPException(
                status_code=400,
                detail="payload_json harus memiliki array: flight_manifest, uld, dan mawb.",
            )

        from openpyxl import Workbook

        wb = Workbook()
        ws_flight = wb.active
        ws_flight.title = "flight_manifest"
        ws_uld = wb.create_sheet("uld")
        ws_mawb = wb.create_sheet("mawb")

        ws_flight.append(FLIGHT_HEADERS)
        ws_uld.append(ULD_HEADERS)
        ws_mawb.append(MAWB_HEADERS)

        def _append_rows(ws, headers: list[str], rows: list):
            for row in rows:
                if not isinstance(row, dict):
                    continue
                ws.append([row.get(header) for header in headers])

        _append_rows(ws_flight, FLIGHT_HEADERS, flight_rows)
        _append_rows(ws_uld, ULD_HEADERS, uld_rows)
        _append_rows(ws_mawb, MAWB_HEADERS, mawb_rows)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def _build_number_build_up(
        flight_date: date,
        airline_code: str | None,
        flight_number: str | None,
        sequence: int,
    ) -> str:
        date_part = flight_date.strftime("%d%m%Y")
        flight_part = "".join(
            ch for ch in f"{airline_code or ''}{flight_number or ''}" if ch.isalnum()
        ).upper()[:12]
        suffix = datetime.now(timezone.utc).strftime("%H%M%S%f")[-6:]
        return f"BL{date_part}{flight_part}{sequence:02d}{suffix}"

    @staticmethod
    def submit_manifest(payload_json: str, db: Session) -> dict:  # noqa: PLR0912, PLR0915
        contents = BuildupService._build_workbook_from_payload(payload_json)
        wb = load_workbook(filename=BytesIO(contents), data_only=True)

        if "flight_manifest" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail="Sheet flight_manifest tidak ditemukan.")
        if "uld" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail="Sheet uld tidak ditemukan.")
        if "mawb" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail="Sheet mawb tidak ditemukan.")

        ws_flight = wb["flight_manifest"]
        ws_uld = wb["uld"]
        ws_mawb = wb["mawb"]

        flight_headers = _read_headers_any(
            ws_flight,
            [FLIGHT_HEADERS, FLIGHT_HEADERS_WITHOUT_TOTAL_VOLUME],
            "flight_manifest",
        )
        _read_headers(ws_uld, ULD_HEADERS, "uld")
        mawb_headers = _read_headers_any(
            ws_mawb,
            [MAWB_HEADERS, MAWB_HEADERS_WITHOUT_TOTAL_WEIGHT, MAWB_HEADERS_LEGACY],
            "mawb",
        )

        flight_entries: dict[tuple[str, date], dict] = {}
        uld_entries: dict[tuple[str, date, str, str], dict] = {}

        for idx, row in enumerate(ws_flight.iter_rows(min_row=2, values_only=True), start=2):
            if not _row_has_values(row):
                continue

            values = dict(
                zip(flight_headers, list(row) + [None] * len(flight_headers), strict=False)
            )

            airline_code = _normalize_identifier(values.get("airline_code"))
            flight_number = _normalize_identifier(values.get("flight_number"))
            flight_date = _parse_date(values.get("flight_date"), "flight_date", idx)
            point_of_loading = _normalize_identifier(values.get("point_of_loading"))
            point_of_unloading = _normalize_identifier(values.get("point_of_unloading"))

            if not all([airline_code, flight_number, point_of_loading, point_of_unloading]):
                raise HTTPException(
                    status_code=400, detail=f"Data flight tidak lengkap di baris {idx}."
                )

            flight_key = (flight_number, flight_date)
            if flight_key in flight_entries:
                raise HTTPException(status_code=400, detail=f"Flight duplikat di baris {idx}.")

            total_pieces = _to_int(values.get("total_pieces"), "total_pieces", idx)
            total_weight = _to_decimal(values.get("total_weight_kg"), "total_weight_kg", idx)
            total_volume = _to_decimal(values.get("total_volume"), "total_volume", idx)

            flight_entries[flight_key] = {
                "airline_code": airline_code,
                "flight_number": flight_number,
                "flight_date": flight_date,
                "aircraft_registration": _normalize_identifier(values.get("aircraft_registration")),
                "point_of_loading": point_of_loading,
                "point_of_unloading": point_of_unloading,
                "total_pieces": total_pieces,
                "total_weight": total_weight,
                "total_volume": total_volume,
                "calculated_total_volume": Decimal("0"),
                "source_document": _normalize_identifier(values.get("source_document"))
                or "manual-form",
                "raw_text": _normalize_identifier(values.get("raw_text")),
                "ulds": [],
            }

        if not flight_entries:
            raise HTTPException(status_code=400, detail="Sheet flight_manifest kosong.")

        for idx, row in enumerate(ws_uld.iter_rows(min_row=2, values_only=True), start=2):
            if not _row_has_values(row):
                continue

            values = dict(zip(ULD_HEADERS, list(row) + [None] * len(ULD_HEADERS), strict=False))
            flight_number = _normalize_identifier(values.get("flight_number"))
            flight_date = _parse_date(values.get("flight_date"), "flight_date", idx)
            uld_type = _normalize_identifier(values.get("uld_type"))
            uld_number = _normalize_identifier(values.get("uld_number"))
            destination = _normalize_identifier(values.get("destination"))

            if not all([flight_number, uld_type, uld_number, destination]):
                raise HTTPException(
                    status_code=400, detail=f"Data ULD tidak lengkap di baris {idx}."
                )

            flight_key = (flight_number, flight_date)
            flight_entry = flight_entries.get(flight_key)
            if not flight_entry:
                raise HTTPException(
                    status_code=400,
                    detail=f"Flight tidak ditemukan untuk ULD di baris {idx}.",
                )

            uld_key = (flight_number, flight_date, uld_type, uld_number)
            if uld_key in uld_entries:
                raise HTTPException(status_code=400, detail=f"ULD duplikat di baris {idx}.")

            uld_payload = {
                "uld_type": uld_type,
                "uld_number": uld_number,
                "uld_owner": _normalize_identifier(values.get("uld_owner")) or "FX",
                "destination": destination,
                "remarks": _normalize_identifier(values.get("remarks")),
                "mawbs": [],
            }

            flight_entry["ulds"].append(uld_payload)
            uld_entries[uld_key] = uld_payload

        supports_total_pieces = "total_pieces" in mawb_headers
        supports_total_weight = "total_weight_kg" in mawb_headers
        mawb_dead_stock_tracker: dict[str, dict] = {}

        for idx, row in enumerate(ws_mawb.iter_rows(min_row=2, values_only=True), start=2):
            if not _row_has_values(row):
                continue

            values = dict(zip(mawb_headers, list(row) + [None] * len(mawb_headers), strict=False))
            flight_number = _normalize_identifier(values.get("flight_number"))
            flight_date = _parse_date(values.get("flight_date"), "flight_date", idx)
            uld_type = _normalize_identifier(values.get("uld_type"))
            uld_number = _normalize_identifier(values.get("uld_number"))
            mawb_prefix = _normalize_identifier(values.get("mawb_prefix"), pad=3)
            mawb_number = _normalize_identifier(values.get("mawb_number"))

            if not all([flight_number, uld_type, uld_number, mawb_prefix, mawb_number]):
                raise HTTPException(
                    status_code=400, detail=f"Data MAWB tidak lengkap di baris {idx}."
                )

            uld_key = (flight_number, flight_date, uld_type, uld_number)
            uld_entry = uld_entries.get(uld_key)
            if not uld_entry:
                raise HTTPException(
                    status_code=400,
                    detail=f"ULD tidak ditemukan untuk MAWB di baris {idx}.",
                )

            pieces = _to_int(values.get("pieces"), "pieces", idx)
            weight = _to_decimal(values.get("weight_kg"), "weight_kg", idx)
            volume = _to_decimal(values.get("volume"), "volume", idx)
            if volume <= 0:
                volume = Decimal("1")
            total_pieces = (
                _to_int(values.get("total_pieces"), "total_pieces", idx, default=pieces)
                if supports_total_pieces
                else None
            )
            total_weight = (
                _to_decimal(values.get("total_weight_kg"), "total_weight_kg", idx, default=weight)
                if supports_total_weight
                else None
            )
            transit_flag = _to_bool(values.get("transit_flag"))
            mawb_full_code = f"{mawb_prefix}-{mawb_number}"
            mawb_key = _normalize_awb_key(mawb_full_code)

            if mawb_key not in mawb_dead_stock_tracker:
                mawb_dead_stock_tracker[mawb_key] = {
                    "mawb": mawb_full_code,
                    "expected_pieces": total_pieces,
                    "expected_weight": total_weight,
                    "input_pieces": 0,
                    "input_weight": Decimal("0"),
                    "build_up_detail_id": None,
                }
            else:
                tracker = mawb_dead_stock_tracker[mawb_key]
                if total_pieces is not None:
                    existing_expected_pieces = tracker.get("expected_pieces")
                    tracker["expected_pieces"] = (
                        max(existing_expected_pieces, total_pieces)
                        if existing_expected_pieces is not None
                        else total_pieces
                    )
                if total_weight is not None:
                    existing_expected_weight = tracker.get("expected_weight")
                    tracker["expected_weight"] = (
                        max(existing_expected_weight, total_weight)
                        if existing_expected_weight is not None
                        else total_weight
                    )

            mawb_dead_stock_tracker[mawb_key]["input_pieces"] += pieces
            mawb_dead_stock_tracker[mawb_key]["input_weight"] += weight
            flight_entry = flight_entries.get((flight_number, flight_date))
            if flight_entry:
                flight_entry["calculated_total_volume"] += volume

            uld_entry["mawbs"].append(
                {
                    "mawb_prefix": mawb_prefix,
                    "mawb_number": mawb_number,
                    "pieces": pieces,
                    "total_pieces": total_pieces,
                    "total_weight_kg": _format_decimal(total_weight)
                    if total_weight is not None
                    else None,
                    "weight_kg": _format_decimal(weight),
                    "volume": _format_decimal(volume),
                    "nature_of_goods": _normalize_identifier(values.get("nature_of_goods")),
                    "route": _normalize_identifier(values.get("route")),
                    "transit_flag": "TRANSIT" if transit_flag else "",
                }
            )

        flight_items = list(flight_entries.values())
        for entry in flight_items:
            if entry["total_volume"] <= 0:
                if entry["calculated_total_volume"] > 0:
                    entry["total_volume"] = entry["calculated_total_volume"]
                else:
                    entry["total_volume"] = Decimal("1")

        flights_payload: list[dict] = []
        for entry in flight_items:
            flights_payload.append(
                {
                    "airline_code": entry["airline_code"],
                    "flight_number": entry["flight_number"],
                    "flight_date": entry["flight_date"].strftime("%Y-%m-%d"),
                    "aircraft_registration": entry["aircraft_registration"],
                    "point_of_loading": entry["point_of_loading"],
                    "point_of_unloading": entry["point_of_unloading"],
                    "total_pieces": entry["total_pieces"],
                    "total_weight_kg": _format_decimal(entry["total_weight"]),
                    "total_volume": _format_decimal(entry["total_volume"]),
                    "source_document": entry["source_document"],
                    "raw_text": entry["raw_text"],
                    "ulds": entry["ulds"],
                }
            )

        pdf_bytes = BuildupService._generate_pdf(flights_payload)
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
        flight_tag = flight_items[0]["flight_number"] if flight_items else "manifest"
        flight_tag = "".join(ch for ch in str(flight_tag) if ch.isalnum() or ch in ("-", "_"))
        if not flight_tag:
            flight_tag = "manifest"

        pdf_filename = f"buildup_manifest_{flight_tag}_{timestamp}.pdf"
        pdf_path = Path(PDF_DIR) / pdf_filename
        pdf_url = f"/pdf/{pdf_filename}"

        header_count = 0
        detail_count = 0
        try:
            PDF_DIR.mkdir(parents=True, exist_ok=True)
            pdf_path.write_bytes(pdf_bytes)

            for index, entry in enumerate(flight_items, start=1):
                official_use = next(
                    (
                        mawb.get("route")
                        for uld in entry["ulds"]
                        for mawb in uld["mawbs"]
                        if mawb.get("route")
                    ),
                    None,
                )

                header_obj = BuildUpHeader(
                    number_build_up=BuildupService._build_number_build_up(
                        flight_date=entry["flight_date"],
                        airline_code=entry["airline_code"],
                        flight_number=entry["flight_number"],
                        sequence=index,
                    ),
                    airlines_code=entry["airline_code"],
                    origin=entry["point_of_loading"],
                    dest=entry["point_of_unloading"],
                    flight_date=entry["flight_date"],
                    for_official_use=official_use,
                    total_pieces=entry["total_pieces"],
                    total_weight=float(entry["total_weight"]),
                    total_volume=float(entry["total_volume"]),
                    pdf_link=pdf_url,
                )
                db.add(header_obj)
                db.flush()
                header_count += 1

                for uld in entry["ulds"]:
                    if uld["mawbs"]:
                        for mawb in uld["mawbs"]:
                            mawb_full_code = f"{mawb['mawb_prefix']}-{mawb['mawb_number']}"
                            detail_obj = BuildUpDetail(
                                header_id=header_obj.id,
                                mawb=mawb_full_code,
                                uld_number=uld["uld_number"],
                                uld_type=uld["uld_type"],
                                pieces=mawb["pieces"],
                                weight=float(mawb["weight_kg"]),
                                volume=float(mawb["volume"]),
                                nature_of_goods=mawb.get("nature_of_goods"),
                                remark=_compose_detail_remark(
                                    base_remark=uld.get("remarks"),
                                    transit_marker=mawb.get("transit_flag"),
                                ),
                            )
                            db.add(detail_obj)
                            db.flush()
                            mawb_tracker = mawb_dead_stock_tracker.get(
                                _normalize_awb_key(mawb_full_code)
                            )
                            if (
                                mawb_tracker
                                and not mawb_tracker.get("build_up_detail_id")
                                and detail_obj.id is not None
                            ):
                                mawb_tracker["build_up_detail_id"] = detail_obj.id
                            detail_count += 1
                    else:
                        detail_obj = BuildUpDetail(
                            header_id=header_obj.id,
                            mawb=None,
                            uld_number=uld["uld_number"],
                            uld_type=uld["uld_type"],
                            pieces=None,
                            weight=None,
                            nature_of_goods=None,
                            remark=uld.get("remarks"),
                        )
                        db.add(detail_obj)
                        db.flush()
                        detail_count += 1

            try:
                with db.begin_nested():
                    tracked_mawbs = [tracker["mawb"] for tracker in mawb_dead_stock_tracker.values()]
                    existing_dead_stocks = []
                    if tracked_mawbs:
                        existing_dead_stocks = (
                            db.query(BuildUpDeadStock)
                            .filter(BuildUpDeadStock.mawb.in_(tracked_mawbs))
                            .order_by(BuildUpDeadStock.id.desc())
                            .all()
                        )
                    latest_dead_stock_by_mawb: dict[str, BuildUpDeadStock] = {}
                    for stock in existing_dead_stocks:
                        key = _normalize_awb_key(stock.mawb)
                        if key and key not in latest_dead_stock_by_mawb:
                            latest_dead_stock_by_mawb[key] = stock

                    for mawb_key, tracker in mawb_dead_stock_tracker.items():
                        expected_pieces = tracker.get("expected_pieces")
                        expected_weight = tracker.get("expected_weight")
                        input_pieces = tracker.get("input_pieces", 0)
                        input_weight = tracker.get("input_weight", Decimal("0"))
                        detail_id = tracker.get("build_up_detail_id")
                        if expected_pieces is None and expected_weight is None:
                            continue

                        remaining_pieces = None
                        if expected_pieces is not None:
                            remaining_pieces = max(int(expected_pieces) - int(input_pieces), 0)

                        remaining_weight = None
                        if expected_weight is not None:
                            remaining_weight = expected_weight - input_weight
                            if remaining_weight < 0:
                                remaining_weight = Decimal("0")

                        has_remaining = (
                            (remaining_pieces is not None and remaining_pieces > 0)
                            or (remaining_weight is not None and remaining_weight > 0)
                        )
                        existing_stock = latest_dead_stock_by_mawb.get(mawb_key)

                        if has_remaining:
                            if existing_stock:
                                existing_stock.build_up_detail_id = detail_id
                                existing_stock.pieces = remaining_pieces
                                existing_stock.weight = (
                                    float(remaining_weight) if remaining_weight is not None else None
                                )
                            else:
                                db.add(
                                    BuildUpDeadStock(
                                        build_up_detail_id=detail_id,
                                        mawb=tracker["mawb"],
                                        pieces=remaining_pieces,
                                        weight=float(remaining_weight)
                                        if remaining_weight is not None
                                        else None,
                                    )
                                )
                        elif existing_stock:
                            existing_stock.build_up_detail_id = detail_id
                            existing_stock.pieces = None
                            existing_stock.weight = None
            except ProgrammingError as exc:
                if _is_missing_dead_stock_table_error(exc):
                    logger.warning(
                        "Tabel build_up_dead_stock belum tersedia. Sinkronisasi dead stock dilewati."
                    )
                else:
                    raise

            db.commit()
        except IntegrityError as exc:
            db.rollback()
            if pdf_path.exists():
                try:
                    pdf_path.unlink()
                except OSError:
                    logger.warning("Gagal menghapus file PDF yang sudah dibuat: %s", pdf_path)
            logger.exception("Gagal menyimpan data build up", exc_info=exc)
            raise HTTPException(status_code=500, detail="Gagal menyimpan data build up.") from exc
        except Exception as exc:
            db.rollback()
            if pdf_path.exists():
                try:
                    pdf_path.unlink()
                except OSError:
                    logger.warning("Gagal menghapus file PDF yang sudah dibuat: %s", pdf_path)
            logger.exception("Gagal menyimpan data build up", exc_info=exc)
            raise HTTPException(status_code=500, detail="Gagal menyimpan data build up.") from exc

        return {
            "message": "Submit build up berhasil.",
            "header_count": header_count,
            "detail_count": detail_count,
            "pdf_url": pdf_url,
            "pdf_filename": pdf_filename,
        }

    @staticmethod
    def _generate_pdf(flights: list[dict]) -> bytes:
        templates_dir = Path(__file__).resolve().parent.parent / "templates"
        env = Environment(
            loader=FileSystemLoader(str(templates_dir)),
            autoescape=select_autoescape(["html", "xml"]),
        )
        manifest = BuildupService._build_manifest_payload(flights)
        print(f"ini kenapa flight? {flights}")
        try:
            template = env.get_template("air_cargo_manifest2.html")
            html_content = template.render(
                manifest=manifest,
                flights=flights,
                generated_at=datetime.now(timezone.utc),
                total_flights=len(flights),
            )
        except HTTPException:
            raise
        except Exception as exc:
            logger.exception("Gagal menyiapkan HTML manifest build up.")
            raise HTTPException(status_code=500, detail="Gagal membuat PDF build up.") from exc

        try:
            from weasyprint import HTML

            pdf_bytes = HTML(string=html_content, base_url=str(templates_dir)).write_pdf()
        except (ImportError, OSError) as exc:
            logger.exception("WeasyPrint dependency sistem belum lengkap.")
            raise HTTPException(
                status_code=500,
                detail=(
                    "WeasyPrint belum siap digunakan di server ini. "
                    "Pasang GTK/Pango runtime sesuai panduan WeasyPrint."
                ),
            ) from exc
        except AttributeError as exc:
            logger.exception("Ketidakcocokan dependency WeasyPrint terdeteksi.")
            raise HTTPException(
                status_code=500,
                detail=(
                    "Dependency PDF tidak kompatibel (weasyprint/pydyf). "
                    "Gunakan pydyf versi >=0.10 dan <0.12, lalu install ulang dependency."
                ),
            ) from exc
        except Exception:
            logger.exception("WeasyPrint gagal menghasilkan output PDF build up.")
            raise HTTPException(
                status_code=500,
                detail="Gagal memproses file PDF build up dengan WeasyPrint.",
            )

        if not pdf_bytes:
            raise HTTPException(status_code=500, detail="File PDF build up kosong.")
        return pdf_bytes
